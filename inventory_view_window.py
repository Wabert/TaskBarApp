# inventory_view_window.py
"""
Interactive inventory view window with Excel-like filtering capabilities
This component is designed to be reused throughout the SuiteView project for displaying any tabular data
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from config import Colors, Fonts, Dimensions
from ui_components import CustomDialog
from typing import Any

class InventoryViewWindow(tk.Toplevel):
    """
    Reusable interactive window for viewing any tabular data with Excel-like filtering
    
    Args:
        parent: Parent window
        data: List of dictionaries containing the data to display
        window_config: Configuration dictionary with:
            - title: Window title (default: "Data View")
            - columns: List of column configurations, each with:
                - key: Dictionary key for this column
                - header: Display name for column header
                - width: Column width (default: 100)
                - type: Data type ('text', 'number', 'date') for sorting (default: 'text')
            - on_item_click: Callback function(item) when item is clicked
            - on_item_double_click: Callback function(item) when item is double-clicked
            - show_stats: Whether to show statistics (default: True)
            - allow_export: Whether to allow Excel export (default: True)
            - window_width: Initial window width (default: 1000)
            - window_height: Initial window height (default: 700)
            - additional_info: Dict of additional info to display in header
    """
    
    def __init__(self, parent, data: list[dict[str, Any]], window_config: dict | None = None):
        super().__init__(parent)
        self.parent = parent
        self.original_data = data.copy() if data else []
        self.filtered_data = self.original_data.copy()
        
        # Parse configuration
        config = window_config or {}
        self.window_title = config.get('title', 'Data View')
        self.column_configs = config.get('columns', self._auto_generate_columns())
        self.on_item_click = config.get('on_item_click')
        self.on_item_double_click = config.get('on_item_double_click')
        self.show_stats = config.get('show_stats', True)
        self.allow_export = config.get('allow_export', True)
        self.window_width = config.get('window_width', 1000)
        self.window_height = config.get('window_height', 700)
        self.additional_info = config.get('additional_info', {})
        
        # Extract column information
        self.columns = [col['key'] for col in self.column_configs]
        self.column_headers = {col['key']: col.get('header', col['key']) for col in self.column_configs}
        self.column_widths = {col['key']: col.get('width', 100) for col in self.column_configs}
        self.column_types = {col['key']: col.get('type', 'text') for col in self.column_configs}
        
        # Filter state tracking
        self.active_filters = {}
        self.column_unique_values = {}
        
        # Window setup
        self.overrideredirect(True)
        self.configure(bg=Colors.DARK_GREEN)
        self.attributes('-topmost', True)
        self.geometry(f"{self.window_width}x{self.window_height}")
        
        # Initialize drag variables
        self.is_dragging = False
        self.drag_start_x = 0
        self.drag_start_y = 0
        
        # Main container
        self.main_frame = tk.Frame(self, bg=Colors.DARK_GREEN, relief=tk.RAISED, bd=2)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=3, pady=3)
        
        # Create UI components
        self.create_custom_title_bar()
        self.content_frame = tk.Frame(self.main_frame, bg=Colors.LIGHT_GREEN)
        self.content_frame.pack(fill=tk.BOTH, expand=True)
        
        self.create_header()
        self.create_data_grid()
        self.create_footer()
        
        # Populate and center
        self.populate_grid()
        self.update_stats()
        self.center_window()
        
        # Bind close event
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
    
    def _auto_generate_columns(self) -> list[dict]:
        """Auto-generate column configuration from data"""
        if not self.original_data:
            return []
        
        # Get all unique keys from data
        all_keys = set()
        for item in self.original_data:
            all_keys.update(item.keys())
        
        # Create column config for each key
        columns = []
        for key in sorted(all_keys):
            columns.append({
                'key': key,
                'header': key.replace('_', ' ').title(),
                'width': 150,
                'type': self._guess_column_type(key)
            })
        
        return columns
    
    def _guess_column_type(self, key: str) -> str:
        """Guess column type based on key name and sample data"""
        key_lower = key.lower()
        
        # Check key name patterns
        if any(word in key_lower for word in ['date', 'time', 'created', 'modified', 'updated']):
            return 'date'
        elif any(word in key_lower for word in ['count', 'number', 'size', 'bytes', 'id', 'qty', 'quantity']):
            return 'number'
        
        # Check sample data
        sample_values = []
        for item in self.original_data[:10]:  # Check first 10 items
            if key in item and item[key] is not None:
                sample_values.append(item[key])
        
        if sample_values:
            # Check if all values are numeric
            try:
                for val in sample_values:
                    float(str(val).replace(',', ''))
                return 'number'
            except:
                pass
        
        return 'text'
    
    def create_custom_title_bar(self):
        """Create custom title bar matching other windows' style"""
        self.title_frame = tk.Frame(self.main_frame, bg=Colors.DARK_GREEN, height=25)
        self.title_frame.pack(fill=tk.X)
        self.title_frame.pack_propagate(False)
        
        # Drag handle
        drag_handle = tk.Label(self.title_frame, text="⋮⋮⋮", bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                              font=('Arial', 8), cursor='fleur')
        drag_handle.pack(side=tk.LEFT, padx=3, pady=3)
        
        # Title with icon
        icon = "📊" if "email" in self.window_title.lower() else "📁"
        title_label = tk.Label(self.title_frame, text=f"{icon} {self.window_title}", 
                              bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                              font=Fonts.DIALOG_TITLE, cursor='fleur')
        title_label.pack(side=tk.LEFT, padx=5, pady=3)
        
        # Close button
        close_btn = tk.Label(self.title_frame, text="×", bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                            font=('Arial', 12, 'bold'), cursor='hand2')
        close_btn.pack(side=tk.RIGHT, padx=5)
        close_btn.bind("<Button-1>", lambda e: self.on_closing())
        
        # Bind drag events
        for widget in [self.title_frame, drag_handle, title_label]:
            widget.bind("<Button-1>", self.start_drag)
            widget.bind("<B1-Motion>", self.do_drag)
            widget.bind("<ButtonRelease-1>", self.end_drag)
    
    def create_header(self):
        """Create header with information"""
        header_frame = tk.Frame(self.content_frame, bg=Colors.LIGHT_GREEN, relief=tk.RAISED, bd=1)
        header_frame.pack(fill=tk.X, padx=2, pady=2)
        
        # Title
        title_label = tk.Label(header_frame, text=self.window_title, 
                              bg=Colors.LIGHT_GREEN, fg=Colors.BLACK,
                              font=('Arial', 14, 'bold'))
        title_label.pack(pady=5)
        
        # Additional information
        if self.additional_info:
            info_frame = tk.Frame(header_frame, bg=Colors.LIGHT_GREEN)
            info_frame.pack(pady=5)
            
            col = 0
            for key, value in self.additional_info.items():
                tk.Label(info_frame, text=f"{key}:", bg=Colors.LIGHT_GREEN, 
                        fg=Colors.BLACK, font=Fonts.DIALOG_LABEL).grid(row=0, column=col, sticky='w', padx=5)
                tk.Label(info_frame, text=str(value), bg=Colors.LIGHT_GREEN, 
                        fg=Colors.DARK_GREEN, font=Fonts.DIALOG_LABEL).grid(row=0, column=col+1, sticky='w', padx=5)
                col += 2
        
        # Stats label
        if self.show_stats:
            self.stats_label = tk.Label(header_frame, text="", bg=Colors.LIGHT_GREEN, 
                                       fg=Colors.BLACK, font=Fonts.DIALOG_LABEL)
            self.stats_label.pack(pady=5)
    
    def create_data_grid(self):
        """Create the main data grid with filtering"""
        grid_frame = tk.Frame(self.content_frame, bg=Colors.LIGHT_GREEN, relief=tk.SUNKEN, bd=1)
        grid_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        
        # Create Treeview
        self.tree = ttk.Treeview(grid_frame, show='tree headings')
        self.tree['columns'] = self.columns
        
        # Configure columns
        self.tree.column('#0', width=0, stretch=False)
        self.tree.heading('#0', text='')
        
        for col in self.columns:
            self.tree.column(col, width=self.column_widths.get(col, 100), anchor='w')
            header_text = self.column_headers.get(col, col)
            self.tree.heading(col, text=header_text, command=lambda c=col: self.show_filter_menu(c))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(grid_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(grid_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack components
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        grid_frame.grid_rowconfigure(0, weight=1)
        grid_frame.grid_columnconfigure(0, weight=1)
        
        # Style
        style = ttk.Style()
        style.configure('Treeview', background=Colors.LIGHT_GREEN, 
                       foreground=Colors.BLACK, fieldbackground=Colors.LIGHT_GREEN)
        style.configure('Treeview.Heading', background=Colors.MEDIUM_GREEN,
                       foreground=Colors.BLACK, font=Fonts.MENU_HEADER)
        
        # Bind click events
        if self.on_item_click:
            self.tree.bind('<ButtonRelease-1>', self._handle_item_click)
        if self.on_item_double_click:
            self.tree.bind('<Double-Button-1>', self._handle_item_double_click)
    
    def _handle_item_click(self, event):
        """Handle single click on item with column detection"""
        selection = self.tree.selection()
        if selection and self.on_item_click:
            item_id = selection[0]
            item_index = self.tree.index(item_id)
            
            if 0 <= item_index < len(self.filtered_data):
                # Determine which column was clicked
                column_id = self.tree.identify_column(event.x)
                
                # Convert column id (#1, #2, etc.) to column index
                if column_id:
                    try:
                        col_index = int(column_id.replace('#', '')) - 1
                        if 0 <= col_index < len(self.columns):
                            column_key = self.columns[col_index]
                            # Pass both item and column to callback
                            if hasattr(self.on_item_click, '__code__') and self.on_item_click.__code__.co_argcount > 2:
                                # New style callback with column info
                                self.on_item_click(self.filtered_data[item_index], column_key)
                            else:
                                # Old style callback without column info
                                self.on_item_click(self.filtered_data[item_index])
                    except:
                        # Fallback to just item
                        self.on_item_click(self.filtered_data[item_index])
    
    def _handle_item_double_click(self, event):
        """Handle double click on item"""
        selection = self.tree.selection()
        if selection and self.on_item_double_click:
            item_id = selection[0]
            item_index = self.tree.index(item_id)
            if 0 <= item_index < len(self.filtered_data):
                self.on_item_double_click(self.filtered_data[item_index])
    
    def create_footer(self):
        """Create footer with action buttons and filter status"""
        footer_frame = tk.Frame(self.content_frame, bg=Colors.DARK_GREEN, height=50)
        footer_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=2, pady=2)
        footer_frame.pack_propagate(False)
        
        # Filter status
        filter_frame = tk.Frame(footer_frame, bg=Colors.DARK_GREEN)
        filter_frame.pack(side=tk.LEFT, fill=tk.Y)
        
        self.filter_status_label = tk.Label(filter_frame, text="No filters applied", 
                                           bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                                           font=Fonts.MENU_ITEM)
        self.filter_status_label.pack(side=tk.LEFT, padx=10, pady=5)
        
        # Action buttons
        button_frame = tk.Frame(footer_frame, bg=Colors.DARK_GREEN)
        button_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Clear Filters
        clear_btn = tk.Button(button_frame, text="Clear All Filters", 
                             bg=Colors.MEDIUM_GREEN, fg=Colors.BLACK,
                             relief=tk.RAISED, bd=1, cursor='hand2',
                             font=Fonts.MENU_ITEM, padx=10,
                             command=self.clear_all_filters)
        clear_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Export to Excel (if enabled)
        if self.allow_export:
            export_btn = tk.Button(button_frame, text="Export to Excel", 
                                  bg=Colors.MEDIUM_GREEN, fg=Colors.BLACK,
                                  relief=tk.RAISED, bd=1, cursor='hand2',
                                  font=Fonts.MENU_ITEM, padx=10,
                                  command=self.export_to_excel)
            export_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Close button
        close_btn = tk.Button(button_frame, text="Close", 
                             bg=Colors.INACTIVE_GRAY, fg=Colors.WHITE,
                             relief=tk.RAISED, bd=1, cursor='hand2',
                             font=Fonts.MENU_ITEM, padx=10,
                             command=self.on_closing)
        close_btn.pack(side=tk.LEFT, padx=5, pady=5)
    
    def populate_grid(self):
        """Populate the grid with current filtered data"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Add filtered data
        for item in self.filtered_data:
            values = []
            for col in self.columns:
                value = item.get(col, '')
                # Format based on type
                if self.column_types.get(col) == 'number' and value != '':
                    try:
                        # Format numbers with commas
                        if isinstance(value, (int, float)):
                            value = f"{value:,}"
                    except:
                        pass
                values.append(str(value))
            
            self.tree.insert('', 'end', values=values)
        
        # Calculate unique values
        self.calculate_unique_values()
    
    def calculate_unique_values(self):
        """Calculate unique values for each column from filtered data"""
        self.column_unique_values = {}
        
        for col in self.columns:
            unique_vals = set()
            for item in self.filtered_data:
                val = item.get(col, '')
                if val != '':
                    unique_vals.add(str(val))
            self.column_unique_values[col] = sorted(list(unique_vals))
    
    def show_filter_menu(self, column):
        """Show filter menu for a specific column"""
        available_values = self.get_available_values_for_column(column)
        FilterMenuDialog(self, column, self.column_headers.get(column, column),
                        available_values, 
                        self.active_filters.get(column, set()), 
                        self.apply_filter)
    
    def get_available_values_for_column(self, column):
        """Get all possible values for a column considering OTHER column filters"""
        available_values = set()
        
        temp_filters = self.active_filters.copy()
        if column in temp_filters:
            del temp_filters[column]
        
        for item in self.original_data:
            include_item = True
            
            for filter_col, filter_values in temp_filters.items():
                item_value = str(item.get(filter_col, ''))
                if item_value not in filter_values:
                    include_item = False
                    break
            
            if include_item:
                val = item.get(column, '')
                if val != '':
                    available_values.add(str(val))
        
        return sorted(list(available_values))
    
    def apply_filter(self, column, selected_values):
        """Apply filter to a specific column"""
        if selected_values:
            self.active_filters[column] = set(selected_values)
        else:
            if column in self.active_filters:
                del self.active_filters[column]
        
        self.filter_data()
        self.update_display()
        self.update_filter_status()
        self.update_column_headers()
    
    def filter_data(self):
        """Apply all active filters to the data"""
        self.filtered_data = []
        
        for item in self.original_data:
            include_item = True
            
            for filter_col, filter_values in self.active_filters.items():
                item_value = str(item.get(filter_col, ''))
                if item_value not in filter_values:
                    include_item = False
                    break
            
            if include_item:
                self.filtered_data.append(item)
    
    def update_display(self):
        """Update the grid display with filtered data"""
        self.populate_grid()
        self.update_stats()
    
    def update_stats(self):
        """Update the statistics display"""
        if not self.show_stats or not hasattr(self, 'stats_label'):
            return
            
        total_original = len(self.original_data)
        total_filtered = len(self.filtered_data)
        
        if total_filtered == total_original:
            stats_text = f"Total Items: {total_original:,}"
        else:
            stats_text = f"Showing: {total_filtered:,} of {total_original:,} items"
        
        self.stats_label.config(text=stats_text)
    
    def update_filter_status(self):
        """Update the filter status display"""
        if not self.active_filters:
            self.filter_status_label.config(text="No filters applied")
        else:
            filter_count = len(self.active_filters)
            filter_text = f"{filter_count} filter{'s' if filter_count > 1 else ''} applied"
            self.filter_status_label.config(text=filter_text)
    
    def update_column_headers(self):
        """Update column headers to show filter indicators"""
        for col in self.columns:
            header_text = self.column_headers.get(col, col)
            if col in self.active_filters:
                self.tree.heading(col, text=f"{header_text} ▼")
            else:
                self.tree.heading(col, text=header_text)
    
    def clear_all_filters(self):
        """Clear all active filters"""
        self.active_filters = {}
        self.filtered_data = self.original_data.copy()
        self.update_display()
        self.update_filter_status()
        self.update_column_headers()
    
    def export_to_excel(self):
        """Export the current filtered data to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import tempfile
            import os
            from datetime import datetime
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Export"
            
            # Header
            ws['A1'] = self.window_title
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Total Items: {len(self.filtered_data):,}"
            
            # Column headers
            for col_idx, col_key in enumerate(self.columns, 1):
                cell = ws.cell(row=5, column=col_idx, value=self.column_headers.get(col_key, col_key))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            # Data rows
            for row_idx, item in enumerate(self.filtered_data, 6):
                for col_idx, col_key in enumerate(self.columns, 1):
                    value = item.get(col_key, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save and open
            temp_dir = tempfile.gettempdir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.window_title.replace(' ', '_')}_{timestamp}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            os.startfile(filepath)
            
            messagebox.showinfo("Export Complete", f"Data exported to:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to Excel:\n{str(e)}")
    
    # Drag and window management methods
    def start_drag(self, event):
        """Start dragging the window"""
        self.is_dragging = True
        self.drag_start_x = event.x_root
        self.drag_start_y = event.y_root
        self.title_frame.configure(bg=Colors.HOVER_GREEN)
    
    def do_drag(self, event):
        """Handle drag motion"""
        if not self.is_dragging:
            return
        
        delta_x = event.x_root - self.drag_start_x
        delta_y = event.y_root - self.drag_start_y
        
        new_x = self.winfo_x() + delta_x
        new_y = self.winfo_y() + delta_y
        
        self.geometry(f"+{new_x}+{new_y}")
        
        self.drag_start_x = event.x_root
        self.drag_start_y = event.y_root
    
    def end_drag(self, event):
        """End dragging operation"""
        self.is_dragging = False
        self.title_frame.configure(bg=Colors.DARK_GREEN)
    
    def center_window(self):
        """Center the window on screen"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def on_closing(self):
        """Handle window closing"""
        self.destroy()


class FilterMenuDialog(CustomDialog):
    """Dialog for selecting filter values for a column"""
    
    def __init__(self, parent, column_key, column_header, unique_values, current_selection, apply_callback):
        super().__init__(parent, f"Filter: {column_header}", width=350, height=400)
        
        self.column_key = column_key
        self.column_header = column_header
        self.unique_values = unique_values
        self.current_selection = current_selection.copy()
        self.apply_callback = apply_callback
        self.parent_window = parent
        
        # Check if filter exists
        self.has_existing_filter = column_key in parent.active_filters
        
        # Default to all selected if no current selection
        if not self.current_selection and not self.has_existing_filter:
            self.current_selection = set(unique_values)
        
        self.create_filter_interface()
        self.create_action_buttons()
    
    def create_filter_interface(self):
        """Create the filter selection interface"""
        # Clear Filter button
        if self.has_existing_filter:
            clear_frame = tk.Frame(self.dialog_content, bg=Colors.LIGHT_GREEN)
            clear_frame.pack(fill=tk.X, pady=(0, 10))
            
            clear_filter_btn = tk.Button(clear_frame, text="Clear Filter for This Column", 
                                       bg=Colors.INACTIVE_GRAY, fg=Colors.WHITE,
                                       command=self.clear_column_filter, 
                                       font=Fonts.DIALOG_LABEL,
                                       cursor='hand2', relief=tk.RAISED, bd=1)
            clear_filter_btn.pack(pady=5)
        
        # Search box
        search_frame = tk.Frame(self.dialog_content, bg=Colors.LIGHT_GREEN)
        search_frame.pack(fill=tk.X, pady=5)
        
        tk.Label(search_frame, text="Search:", bg=Colors.LIGHT_GREEN, 
                fg=Colors.BLACK, font=Fonts.DIALOG_LABEL).pack(side=tk.LEFT)
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_list)
        search_entry = tk.Entry(search_frame, textvariable=self.search_var, 
                               font=Fonts.DIALOG_LABEL)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        
        # Select All / None buttons
        select_frame = tk.Frame(self.dialog_content, bg=Colors.LIGHT_GREEN)
        select_frame.pack(fill=tk.X, pady=5)
        
        select_all_btn = tk.Button(select_frame, text="Select All", 
                                  bg=Colors.MEDIUM_GREEN, fg=Colors.BLACK,
                                  command=self.select_all, font=Fonts.DIALOG_LABEL)
        select_all_btn.pack(side=tk.LEFT, padx=5)
        
        select_none_btn = tk.Button(select_frame, text="Select None", 
                                   bg=Colors.MEDIUM_GREEN, fg=Colors.BLACK,
                                   command=self.select_none, font=Fonts.DIALOG_LABEL)
        select_none_btn.pack(side=tk.LEFT, padx=5)
        
        # Listbox with checkboxes
        list_frame = tk.Frame(self.dialog_content, bg=Colors.LIGHT_GREEN)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.filter_tree = ttk.Treeview(list_frame, show='tree', height=12)
        self.filter_tree.column('#0', width=300)
        
        filter_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, 
                                        command=self.filter_tree.yview)
        self.filter_tree.configure(yscrollcommand=filter_scrollbar.set)
        
        self.filter_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        filter_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.populate_filter_list()
        
        # Bind click events
        self.filter_tree.bind('<Button-1>', self.on_click)
        self.filter_tree.bind('<Return>', self.toggle_item)
    
    def clear_column_filter(self):
        """Clear the filter for this specific column"""
        self.apply_callback(self.column_key, [])
        self.destroy()
    
    def on_click(self, event):
        """Handle click on items"""
        item = self.filter_tree.identify('item', event.x, event.y)
        if item:
            self.filter_tree.selection_set(item)
            self.toggle_item()
    
    def populate_filter_list(self, search_text=""):
        """Populate the filter list"""
        for item in self.filter_tree.get_children():
            self.filter_tree.delete(item)
        
        filtered_values = [val for val in self.unique_values 
                          if search_text.lower() in val.lower()] if search_text else self.unique_values
        
        for value in filtered_values:
            checkbox = "☑" if value in self.current_selection else "☐"
            display_text = f"{checkbox} {value}"
            self.filter_tree.insert('', 'end', text=display_text, values=[value])
    
    def filter_list(self, *args):
        """Filter the list based on search"""
        self.populate_filter_list(self.search_var.get())
    
    def toggle_item(self, event=None):
        """Toggle selection of an item"""
        selected_item = self.filter_tree.selection()
        if not selected_item:
            return
        
        item_id = selected_item[0]
        values = self.filter_tree.item(item_id, 'values')
        if values:
            value = values[0]
            
            if value in self.current_selection:
                self.current_selection.remove(value)
            else:
                self.current_selection.add(value)
            
            self.populate_filter_list(self.search_var.get())
    
    def select_all(self):
        """Select all visible items"""
        search_text = self.search_var.get()
        filtered_values = [val for val in self.unique_values 
                          if search_text.lower() in val.lower()] if search_text else self.unique_values
        
        for value in filtered_values:
            self.current_selection.add(value)
        
        self.populate_filter_list(search_text)
    
    def select_none(self):
        """Deselect all visible items"""
        search_text = self.search_var.get()
        filtered_values = [val for val in self.unique_values 
                          if search_text.lower() in val.lower()] if search_text else self.unique_values
        
        for value in filtered_values:
            self.current_selection.discard(value)
        
        self.populate_filter_list(search_text)
    
    def create_action_buttons(self):
        """Create OK and Cancel buttons"""
        button_container = tk.Frame(self.button_frame, bg=Colors.LIGHT_GREEN)
        button_container.pack(expand=True)
        
        ok_btn = tk.Button(button_container, text="OK", 
                          bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                          command=self.apply_filter, width=Dimensions.DIALOG_BUTTON_WIDTH,
                          font=Fonts.DIALOG_BUTTON, relief=tk.RAISED, bd=1)
        ok_btn.pack(side=tk.LEFT, padx=10)
        
        cancel_btn = tk.Button(button_container, text="Cancel", 
                              bg=Colors.INACTIVE_GRAY, fg=Colors.WHITE,
                              command=self.cancel, width=Dimensions.DIALOG_BUTTON_WIDTH,
                              font=Fonts.DIALOG_BUTTON, relief=tk.RAISED, bd=1)
        cancel_btn.pack(side=tk.LEFT, padx=10)
        
        ok_btn.focus_set()
    
    def apply_filter(self):
        """Apply the selected filter"""
        self.apply_callback(self.column_key, list(self.current_selection))
        self.destroy()
    
    def cancel(self):
        """Cancel without applying changes"""
        self.destroy()