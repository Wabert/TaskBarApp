# simple_window_factory.py
"""
Window Factory - Creates custom windows with green styling and advanced functionality
Includes simple windows, inventory view windows, and dialog components
Using inheritance approach for cleaner API
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime
from typing import Any
import json
import os
from config import Colors, Fonts, Dimensions

# Session storage for window positions
_session_window_positions = {}

# Permanent storage file path
_permanent_positions_file = "window_positions.json"

class SimpleWindow(tk.Toplevel):
    def __init__(self, parent, title=None, resize_handles=None, movable=True, location_persistence="none", close_on=None):
        """
        Create a custom window with green styling
        
        Args:
            parent: Parent window (required)
            title: Window title text, or None to hide titlebar completely
            resize_handles: List of sides that can be resized ["left", "right", "top", "bottom"] or None for no resizing
            movable: Whether the window can be moved by dragging (default: True)
            location_persistence: Position persistence mode - "none", "session", or "permanent" (default: "none")
            close_on: List of close methods - ["x_button", "click_outside", "toggle"] (default: ["x_button"])
        """
        # Initialize Toplevel
        super().__init__(parent)
        
        # Remove default title bar
        self.overrideredirect(True)
        
        # Store configuration
        self.title_text = title
        self.resize_handles = resize_handles or []
        self.movable = movable
        self.location_persistence = location_persistence
        self.close_on = close_on or ['x_button']
        
        # For session persistence, use a consistent ID based on title
        # For unique instances, still use memory address
        if location_persistence == "session" and title:
            self.window_id = f"session_{title}"
        else:
            self.window_id = f"{title or 'NoTitle'}_{id(self)}"
        
        # Colors - matching the mockup
        self.border_color = Colors.DARK_GREEN
        self.header_bg = Colors.DARK_GREEN
        self.content_bg = Colors.LIGHT_GREEN
        self.text_color = Colors.WHITE
        self.border_width = 4
        


        # Window state
        self._drag_start_x = 0
        self._drag_start_y = 0
        self._resize_start_x = 0
        self._resize_start_y = 0
        self._resize_start_width = 0
        self._resize_start_height = 0
        self._resizing = False
        self._resize_side = None
        
        # Toggle functionality
        self._toggle_controls = []  # List of controls that can toggle this window
        self._original_commands = {}  # Store original commands to restore later
        
        # Build the window
        self._create_window()
        
        # Load saved position if persistence is enabled
        if self.movable and self.location_persistence != "none":
            self._load_position()
        
        # Set default size if no saved data exists
        if self.location_persistence != "none":
            self.after(1, self._set_default_size_if_needed)
        
    def _create_window(self):
        """Build the window structure"""
        # Configure window background
        self.configure(bg=self.content_bg)
        
        # Main container with border
        self.main_frame = tk.Frame(self, bg=self.border_color)
        self.main_frame.pack(fill="both", expand=True)
        
        # Inner container
        self.inner_frame = tk.Frame(self.main_frame, bg=self.content_bg)
        self.inner_frame.place(x=self.border_width, y=self.border_width, 
                              relwidth=1, relheight=1, 
                              width=-2*self.border_width, 
                              height=-2*self.border_width)
        
        # Header bar (only if title is not None)
        if self.title_text is not None:
            self.header_frame = tk.Frame(self.inner_frame, bg=self.header_bg, height=32)
            self.header_frame.pack(fill="x", side="top")
            self.header_frame.pack_propagate(False)
            
            # Title label
            self.title_label = tk.Label(self.header_frame, text=self.title_text, 
                                       bg=self.header_bg, fg=self.text_color,
                                       font=("Arial", 10), anchor="w")
            self.title_label.pack(side="left", padx=8, fill="y")
            
            # Close button (only if 'x_button' is in close_on)
            if 'x_button' in self.close_on:
                self.close_button = tk.Button(self.header_frame, text="✕", 
                                             bg=self.header_bg, fg=self.text_color,
                                             font=("Arial", 12), bd=0,
                                             activebackground=self.header_bg,
                                             activeforeground=self.text_color,
                                             command=self.close_window)
                self.close_button.pack(side="right", padx=8)
        
        # Content area
        self.content_frame = tk.Frame(self.inner_frame, bg=Colors.WHITE)
        self.content_frame.pack(fill="both", expand=True)
        
        # Bind window dragging (if movable)
        if self.movable:
            if self.title_text is not None:
                # Bind to header elements if header exists
                self.header_frame.bind("<Button-1>", self._start_drag)
                self.header_frame.bind("<B1-Motion>", self._drag_window)
                self.title_label.bind("<Button-1>", self._start_drag)
                self.title_label.bind("<B1-Motion>", self._drag_window)
            else:
                # Bind to content frame if no header (but avoid conflicts with content)
                self.content_frame.bind("<Button-1>", self._start_drag)
                self.content_frame.bind("<B1-Motion>", self._drag_window)
        
        # Set up resize bindings
        self._setup_resize_bindings()
        
        # Set up click outside close if requested
        if 'click_outside' in self.close_on:
            self._setup_click_outside_close()
        
        # Set up toggle functionality if requested
        if 'toggle' in self.close_on:
            self._setup_toggle_close()
        
    def _setup_resize_bindings(self):
        """Set up mouse bindings for resizing"""
        if not self.resize_handles:
            return
            
        # Bind mouse motion to detect when over resize areas
        self.bind("<Motion>", self._on_mouse_motion)
        self.bind("<Button-1>", self._start_resize)
        self.bind("<B1-Motion>", self._do_resize)
        self.bind("<ButtonRelease-1>", self._stop_resize)
        
    def _on_mouse_motion(self, event):
        """Change cursor when over resize areas"""
        if self._resizing:
            return
            
        x, y = event.x, event.y
        width, height = self.winfo_width(), self.winfo_height()
        
        cursor = ""
        if "left" in self.resize_handles and x < self.border_width:
            cursor = "sb_h_double_arrow"
        elif "right" in self.resize_handles and x > width - self.border_width:
            cursor = "sb_h_double_arrow"
        elif "top" in self.resize_handles and y < self.border_width:
            cursor = "sb_v_double_arrow"
        elif "bottom" in self.resize_handles and y > height - self.border_width:
            cursor = "sb_v_double_arrow"
        
        self.config(cursor=cursor)
        
    def _start_resize(self, event):
        """Start resizing if clicked on a resize area"""
        x, y = event.x, event.y
        width, height = self.winfo_width(), self.winfo_height()
        
        self._resize_side = None
        if "left" in self.resize_handles and x < self.border_width:
            self._resize_side = "left"
        elif "right" in self.resize_handles and x > width - self.border_width:
            self._resize_side = "right"
        elif "top" in self.resize_handles and y < self.border_width:
            self._resize_side = "top"
        elif "bottom" in self.resize_handles and y > height - self.border_width:
            self._resize_side = "bottom"
            
        if self._resize_side:
            self._resizing = True
            self._resize_start_x = event.x_root
            self._resize_start_y = event.y_root
            self._resize_start_width = width
            self._resize_start_height = height
            self._resize_start_left = self.winfo_x()
            self._resize_start_top = self.winfo_y()
            
    def _do_resize(self, event):
        """Perform the resize"""
        if not self._resizing or not self._resize_side:
            return
            
        dx = event.x_root - self._resize_start_x
        dy = event.y_root - self._resize_start_y
        
        if self._resize_side == "left":
            new_width = self._resize_start_width - dx
            new_x = self._resize_start_left + dx
            if new_width > 100:  # Minimum width
                self.geometry(f"{new_width}x{self._resize_start_height}+{new_x}+{self._resize_start_top}")
                
        elif self._resize_side == "right":
            new_width = self._resize_start_width + dx
            if new_width > 100:
                self.geometry(f"{new_width}x{self._resize_start_height}")
                
        elif self._resize_side == "top":
            new_height = self._resize_start_height - dy
            new_y = self._resize_start_top + dy
            if new_height > 100:  # Minimum height
                self.geometry(f"{self._resize_start_width}x{new_height}+{self._resize_start_left}+{new_y}")
                
        elif self._resize_side == "bottom":
            new_height = self._resize_start_height + dy
            if new_height > 100:
                self.geometry(f"{self._resize_start_width}x{new_height}")
                
    def _stop_resize(self, event):
        """Stop resizing"""
        self._resizing = False
        self._resize_side = None
        
        # Save position and size if persistence is enabled
        if self.location_persistence != "none":
            self._save_position()
        
    def _start_drag(self, event):
        """Start dragging the window"""
        self._drag_start_x = event.x_root
        self._drag_start_y = event.y_root
        
    def _drag_window(self, event):
        """Drag the window"""
        x = self.winfo_x() + (event.x_root - self._drag_start_x)
        y = self.winfo_y() + (event.y_root - self._drag_start_y)
        self.geometry(f"+{x}+{y}")
        self._drag_start_x = event.x_root
        self._drag_start_y = event.y_root
        
        # Save position if persistence is enabled
        if self.location_persistence != "none":
            self._save_position()
        
    def close_window(self):
        """Close the window"""
        # Clean up toggle controls
        if hasattr(self, '_toggle_controls'):
            for control in self._toggle_controls.copy():
                self.unregister_toggle_control(control)
        
        self.destroy()
        
    def get_content_frame(self):
        """Return the content frame for adding widgets"""
        return self.content_frame
    
    def _save_position(self):
        """Save the current window position and size based on persistence mode"""
        if self.location_persistence == "none":
            return
        
        try:
            x = self.winfo_x()
            y = self.winfo_y()
            width = self.winfo_width()
            height = self.winfo_height()
            position_data = {"x": x, "y": y, "width": width, "height": height}
            
            if self.location_persistence == "session":
                _session_window_positions[self.window_id] = position_data
            elif self.location_persistence == "permanent":
                self._save_permanent_position(position_data)
        except Exception as e:
            # Silently handle any errors in position saving
            pass
    
    def _load_position(self):
        """Load the saved window position and size based on persistence mode"""
        if self.location_persistence == "none":
            return
        
        try:
            position_data = None
            
            if self.location_persistence == "session":
                position_data = _session_window_positions.get(self.window_id)
            elif self.location_persistence == "permanent":
                position_data = self._load_permanent_position()
            
            if position_data and "x" in position_data and "y" in position_data:
                # Validate position is within screen bounds
                screen_width = self.winfo_screenwidth()
                screen_height = self.winfo_screenheight()
                
                x = max(0, min(position_data["x"], screen_width - 100))  # Ensure window is visible
                y = max(0, min(position_data["y"], screen_height - 100))
                
                # Get saved dimensions if available
                if "width" in position_data and "height" in position_data:
                    width = max(100, position_data["width"])  # Minimum width
                    height = max(100, position_data["height"])  # Minimum height
                    self.geometry(f"{width}x{height}+{x}+{y}")
                else:
                    # Only position was saved (backward compatibility)
                    self.geometry(f"+{x}+{y}")
        except Exception as e:
            # Silently handle any errors in position loading
            pass
    
    def _save_permanent_position(self, position_data):
        """Save position and size to permanent storage"""
        try:
            # Load existing positions
            positions = {}
            if os.path.exists(_permanent_positions_file):
                with open(_permanent_positions_file, 'r') as f:
                    positions = json.load(f)
            
            # Update position and size for this window
            positions[self.title_text] = position_data
            
            # Save back to file
            with open(_permanent_positions_file, 'w') as f:
                json.dump(positions, f, indent=2)
        except Exception as e:
            # Silently handle any errors in permanent storage
            pass
    
    def _load_permanent_position(self):
        """Load position and size from permanent storage"""
        try:
            if os.path.exists(_permanent_positions_file):
                with open(_permanent_positions_file, 'r') as f:
                    positions = json.load(f)
                return positions.get(self.title_text)
        except Exception as e:
            # Silently handle any errors in permanent storage
            pass
        return None
    
    def _set_default_size_if_needed(self):
        """Set default size and position if no saved data was applied"""
        try:
            # Check if window has a reasonable size (indicating saved data was applied)
            current_width = self.winfo_width()
            current_height = self.winfo_height()
            
            # If window is very small (default tkinter size), set our defaults
            if current_width <= 200 and current_height <= 200:
                self.geometry("400x300+100+100")
        except Exception as e:
            # If there's any error, set default size anyway
            try:
                self.geometry("400x300+100+100")
            except:
                pass

    def _setup_click_outside_close(self):
        """Set up click outside to close functionality"""
        # Bind to focus out event
        self.bind('<FocusOut>', self._on_focus_out)
        
    def _on_focus_out(self, event):
        """Handle focus out event for click outside close"""
        # Schedule a check to see if we should close
        self.after(50, self._check_if_should_close)
        
    def _check_if_should_close(self):
        """Check if we should close due to focus change"""
        try:
            # Get the currently focused widget
            focused = self.focus_get()
            
            # If no widget has focus, close the window
            if focused is None:
                self.close_window()
                return
                
            # Check if the focused widget is within our window hierarchy
            widget = focused
            while widget:
                if widget == self:
                    # Focus is within our window, don't close
                    return
                widget = widget.master
            
            # Focus is outside our window, close it
            self.close_window()
            
        except:
            # If there's any error, don't close
            pass
    
    def _setup_toggle_close(self):
        """Set up toggle functionality - this will be called when toggle controls are registered"""
        pass  # Implementation happens in register_toggle_control
    
    def register_toggle_control(self, control):
        """
        Register a control (button, label, etc.) that can toggle this window open/closed
        
        Args:
            control: The tkinter widget that should toggle this window
        """
        if control not in self._toggle_controls:
            self._toggle_controls.append(control)
            
            # Store the original command if it exists
            if hasattr(control, 'cget') and control.cget('command'):
                self._original_commands[control] = control.cget('command')
            
            # Set up the toggle command
            if hasattr(control, 'configure'):
                control.configure(command=self._toggle_window)
            elif hasattr(control, 'bind'):
                # For labels or other widgets without command, use click binding
                control.bind("<Button-1>", lambda e: self._toggle_window())
    
    def _toggle_window(self):
        """Toggle the window open/closed"""
        try:
            if self.winfo_exists() and self.winfo_viewable():
                # Window is visible, close it
                self.close_window()
            else:
                # Window is not visible, show it
                self.lift()
                self.focus_force()
        except:
            # If there's an error checking state, try to show the window
            try:
                self.lift()
                self.focus_force()
            except:
                pass
    
    def unregister_toggle_control(self, control):
        """
        Remove a control from toggle functionality and restore its original command
        
        Args:
            control: The tkinter widget to remove from toggle functionality
        """
        if control in self._toggle_controls:
            self._toggle_controls.remove(control)
            
            # Restore original command if it existed
            if control in self._original_commands:
                if hasattr(control, 'configure'):
                    control.configure(command=self._original_commands[control])
                del self._original_commands[control]
            else:
                # No original command, set to None
                if hasattr(control, 'configure'):
                    control.configure(command=None)

class InventoryViewWindow(SimpleWindow):
    """
    Reusable interactive window for viewing any tabular data with Excel-like filtering
    Inherits from SimpleWindow for consistent styling and behavior
    
    Args:
        parent: Parent window
        data: List of dictionaries containing the data to display
        window_config: Configuration dictionary with:
            - title: Window title (default: "Data View")
            - columns: List of column configurations, each with:
                - key: Dictionary key for this column
                - header: Display name for column header
                - width: Column width (default: 100)
                - type: Data type ('text', 'number', 'date') for sorting (default: 'text')
            - on_item_click: Callback function(item) when item is clicked
            - on_item_double_click: Callback function(item) when item is double-clicked
            - show_stats: Whether to show statistics (default: True)
            - allow_export: Whether to allow Excel export (default: True)
            - window_width: Initial window width (default: 1000)
            - window_height: Initial window height (default: 700)
            - additional_info: Dict of additional info to display in header
    """
    
    def __init__(self, parent, data: list[dict[str, Any]], window_config: dict | None = None):
        # Parse configuration
        config = window_config or {}
        self.window_title = config.get('title', 'Data View')
        self.window_width = config.get('window_width', 1000)
        self.window_height = config.get('window_height', 700)
        
        # Initialize SimpleWindow with resize handles
        super().__init__(parent, self.window_title, resize_handles=["left", "right", "bottom"], 
                        location_persistence=config.get('location_persistence', 'none'))
        
        # Set window size
        self.geometry(f"{self.window_width}x{self.window_height}")
        
        # Store data
        self.original_data = data.copy() if data else []
        self.filtered_data = self.original_data.copy()
        
        # Configuration
        self.column_configs = config.get('columns', self._auto_generate_columns())
        self.on_item_click = config.get('on_item_click')
        self.on_item_double_click = config.get('on_item_double_click')
        self.show_stats = config.get('show_stats', True)
        self.allow_export = config.get('allow_export', True)
        self.additional_info = config.get('additional_info', {})
        
        # Extract column information
        self.columns = [col['key'] for col in self.column_configs]
        self.column_headers = {col['key']: col.get('header', col['key']) for col in self.column_configs}
        self.column_widths = {col['key']: col.get('width', 100) for col in self.column_configs}
        self.column_types = {col['key']: col.get('type', 'text') for col in self.column_configs}
        
        # Filter state tracking
        self.active_filters = {}
        self.column_unique_values = {}
        
        # Create UI components in the content_frame from SimpleWindow
        self.create_header()
        self.create_data_grid()
        self.create_footer()
        
        # Populate and center
        self.populate_grid()
        self.update_stats()
        self.center_window()
        
        # Make topmost
        self.attributes('-topmost', True)
        
    def _auto_generate_columns(self) -> list[dict]:
        """Auto-generate column configuration from data"""
        if not self.original_data:
            return []
        
        # Get all unique keys from data
        all_keys = set()
        for item in self.original_data:
            all_keys.update(item.keys())
        
        # Create column config for each key
        columns = []
        for key in sorted(all_keys):
            columns.append({
                'key': key,
                'header': key.replace('_', ' ').title(),
                'width': 150,
                'type': self._guess_column_type(key)
            })
        
        return columns
    
    def _guess_column_type(self, key: str) -> str:
        """Guess column type based on key name and sample data"""
        key_lower = key.lower()
        
        # Check key name patterns
        if any(word in key_lower for word in ['date', 'time', 'created', 'modified', 'updated']):
            return 'date'
        elif any(word in key_lower for word in ['count', 'number', 'size', 'bytes', 'id', 'qty', 'quantity']):
            return 'number'
        
        # Check sample data
        sample_values = []
        for item in self.original_data[:10]:  # Check first 10 items
            if key in item and item[key] is not None:
                sample_values.append(item[key])
        
        if sample_values:
            # Check if all values are numeric
            try:
                for val in sample_values:
                    float(str(val).replace(',', ''))
                return 'number'
            except:
                pass
        
        return 'text'
    
    def create_header(self):
        """Create header with information"""
        header_frame = tk.Frame(self.content_frame, bg=Colors.LIGHT_GREEN, relief=tk.RAISED, bd=1)
        header_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Additional information
        if self.additional_info:
            info_frame = tk.Frame(header_frame, bg=Colors.LIGHT_GREEN)
            info_frame.pack(pady=5)
            
            col = 0
            for key, value in self.additional_info.items():
                tk.Label(info_frame, text=f"{key}:", bg=Colors.LIGHT_GREEN, 
                        fg=Colors.BLACK, font=Fonts.DIALOG_LABEL).grid(row=0, column=col, sticky='w', padx=5)
                tk.Label(info_frame, text=str(value), bg=Colors.LIGHT_GREEN, 
                        fg=Colors.DARK_GREEN, font=Fonts.DIALOG_LABEL).grid(row=0, column=col+1, sticky='w', padx=5)
                col += 2
        
        # Stats label
        if self.show_stats:
            self.stats_label = tk.Label(header_frame, text="", bg=Colors.LIGHT_GREEN, 
                                       fg=Colors.BLACK, font=Fonts.DIALOG_LABEL)
            self.stats_label.pack(pady=2)
    
    def create_data_grid(self):
        """Create the main data grid with filtering"""
        grid_frame = tk.Frame(self.content_frame, bg=Colors.LIGHT_GREEN, relief=tk.SUNKEN, bd=1)
        grid_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 5))
        
        # Create Treeview
        self.tree = ttk.Treeview(grid_frame, show='tree headings')
        self.tree['columns'] = self.columns
        
        # Configure columns
        self.tree.column('#0', width=0, stretch=False)
        self.tree.heading('#0', text='')
        
        for col in self.columns:
            self.tree.column(col, width=self.column_widths.get(col, 100), anchor='w')
            header_text = self.column_headers.get(col, col)
            self.tree.heading(col, text=header_text, command=lambda c=col: self.show_filter_menu(c))
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(grid_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(grid_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Pack components
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        grid_frame.grid_rowconfigure(0, weight=1)
        grid_frame.grid_columnconfigure(0, weight=1)
        
        # Style
        style = ttk.Style()
        style.configure('Treeview', background=Colors.LIGHT_GREEN, 
                       foreground=Colors.BLACK, fieldbackground=Colors.LIGHT_GREEN)
        style.configure('Treeview.Heading', background=Colors.MEDIUM_GREEN,
                       foreground=Colors.WHITE, font=Fonts.MENU_HEADER)
        
        # Bind click events
        if self.on_item_click:
            self.tree.bind('<ButtonRelease-1>', self._handle_item_click)
        if self.on_item_double_click:
            self.tree.bind('<Double-Button-1>', self._handle_item_double_click)
    
    def _handle_item_click(self, event):
        """Handle single click on item with column detection"""
        selection = self.tree.selection()
        if selection and self.on_item_click:
            item_id = selection[0]
            item_index = self.tree.index(item_id)
            
            if 0 <= item_index < len(self.filtered_data):
                # Determine which column was clicked
                column_id = self.tree.identify_column(event.x)
                
                # Convert column id (#1, #2, etc.) to column index
                if column_id:
                    try:
                        col_index = int(column_id.replace('#', '')) - 1
                        if 0 <= col_index < len(self.columns):
                            column_key = self.columns[col_index]
                            # Pass both item and column to callback
                            if hasattr(self.on_item_click, '__code__') and self.on_item_click.__code__.co_argcount > 2:
                                # New style callback with column info
                                self.on_item_click(self.filtered_data[item_index], column_key)
                            else:
                                # Old style callback without column info
                                self.on_item_click(self.filtered_data[item_index])
                    except:
                        # Fallback to just item
                        self.on_item_click(self.filtered_data[item_index])
    
    def _handle_item_double_click(self, event):
        """Handle double click on item"""
        selection = self.tree.selection()
        if selection and self.on_item_double_click:
            item_id = selection[0]
            item_index = self.tree.index(item_id)
            if 0 <= item_index < len(self.filtered_data):
                self.on_item_double_click(self.filtered_data[item_index])
    
    def create_footer(self):
        """Create footer with action buttons and filter status"""
        footer_frame = tk.Frame(self.content_frame, bg=Colors.DARK_GREEN, height=40)
        footer_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=5, pady=(0, 5))
        footer_frame.pack_propagate(False)
        
        # Filter status
        filter_frame = tk.Frame(footer_frame, bg=Colors.DARK_GREEN)
        filter_frame.pack(side=tk.LEFT, fill=tk.Y)
        
        self.filter_status_label = tk.Label(filter_frame, text="No filters applied", 
                                           bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                                           font=Fonts.MENU_ITEM)
        self.filter_status_label.pack(side=tk.LEFT, padx=10, pady=5)
        
        # Action buttons
        button_frame = tk.Frame(footer_frame, bg=Colors.DARK_GREEN)
        button_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Clear Filters
        clear_btn = tk.Button(button_frame, text="Clear Filters", 
                             bg=Colors.MEDIUM_GREEN, fg=Colors.WHITE,
                             relief=tk.RAISED, bd=1, cursor='hand2',
                             font=Fonts.MENU_ITEM, padx=10,
                             command=self.clear_all_filters)
        clear_btn.pack(side=tk.LEFT, padx=5, pady=5)
        
        # Export to Excel (if enabled)
        if self.allow_export:
            export_btn = tk.Button(button_frame, text="Export", 
                                  bg=Colors.MEDIUM_GREEN, fg=Colors.WHITE,
                                  relief=tk.RAISED, bd=1, cursor='hand2',
                                  font=Fonts.MENU_ITEM, padx=10,
                                  command=self.export_to_excel)
            export_btn.pack(side=tk.LEFT, padx=5, pady=5)
    
    def populate_grid(self):
        """Populate the grid with current filtered data"""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Add filtered data
        for item in self.filtered_data:
            values = []
            for col in self.columns:
                value = item.get(col, '')
                # Format based on type
                if self.column_types.get(col) == 'number' and value != '':
                    try:
                        # Format numbers with commas
                        if isinstance(value, (int, float)):
                            value = f"{value:,}"
                    except:
                        pass
                values.append(str(value))
            
            self.tree.insert('', 'end', values=values)
        
        # Calculate unique values
        self.calculate_unique_values()
    
    def calculate_unique_values(self):
        """Calculate unique values for each column from filtered data"""
        self.column_unique_values = {}
        
        for col in self.columns:
            unique_vals = set()
            for item in self.filtered_data:
                val = item.get(col, '')
                if val != '':
                    unique_vals.add(str(val))
            self.column_unique_values[col] = sorted(list(unique_vals))
    
    def show_filter_menu(self, column):
        """Show filter menu for a specific column"""
        available_values = self.get_available_values_for_column(column)
        FilterMenuDialog(self, column, self.column_headers.get(column, column),
                        available_values, 
                        self.active_filters.get(column, set()), 
                        self.apply_filter)
    
    def get_available_values_for_column(self, column):
        """Get all possible values for a column considering OTHER column filters"""
        available_values = set()
        
        temp_filters = self.active_filters.copy()
        if column in temp_filters:
            del temp_filters[column]
        
        for item in self.original_data:
            include_item = True
            
            for filter_col, filter_values in temp_filters.items():
                item_value = str(item.get(filter_col, ''))
                if item_value not in filter_values:
                    include_item = False
                    break
            
            if include_item:
                val = item.get(column, '')
                if val != '':
                    available_values.add(str(val))
        
        return sorted(list(available_values))
    
    def apply_filter(self, column, selected_values):
        """Apply filter to a specific column"""
        if selected_values:
            self.active_filters[column] = set(selected_values)
        else:
            if column in self.active_filters:
                del self.active_filters[column]
        
        self.filter_data()
        self.update_display()
        self.update_filter_status()
        self.update_column_headers()
    
    def filter_data(self):
        """Apply all active filters to the data"""
        self.filtered_data = []
        
        for item in self.original_data:
            include_item = True
            
            for filter_col, filter_values in self.active_filters.items():
                item_value = str(item.get(filter_col, ''))
                if item_value not in filter_values:
                    include_item = False
                    break
            
            if include_item:
                self.filtered_data.append(item)
    
    def update_display(self):
        """Update the grid display with filtered data"""
        self.populate_grid()
        self.update_stats()
    
    def update_stats(self):
        """Update the statistics display"""
        if not self.show_stats or not hasattr(self, 'stats_label'):
            return
            
        total_original = len(self.original_data)
        total_filtered = len(self.filtered_data)
        
        if total_filtered == total_original:
            stats_text = f"Total Items: {total_original:,}"
        else:
            stats_text = f"Showing: {total_filtered:,} of {total_original:,} items"
        
        self.stats_label.config(text=stats_text)
    
    def update_filter_status(self):
        """Update the filter status display"""
        if not self.active_filters:
            self.filter_status_label.config(text="No filters applied")
        else:
            filter_count = len(self.active_filters)
            filter_text = f"{filter_count} filter{'s' if filter_count > 1 else ''} applied"
            self.filter_status_label.config(text=filter_text)
    
    def update_column_headers(self):
        """Update column headers to show filter indicators"""
        for col in self.columns:
            header_text = self.column_headers.get(col, col)
            if col in self.active_filters:
                self.tree.heading(col, text=f"{header_text} ▼")
            else:
                self.tree.heading(col, text=header_text)
    
    def clear_all_filters(self):
        """Clear all active filters"""
        self.active_filters = {}
        self.filtered_data = self.original_data.copy()
        self.update_display()
        self.update_filter_status()
        self.update_column_headers()
    
    def export_to_excel(self):
        """Export the current filtered data to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            import tempfile
            import os
            from datetime import datetime
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Export"
            
            # Header
            ws['A1'] = self.window_title
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Total Items: {len(self.filtered_data):,}"
            
            # Column headers
            for col_idx, col_key in enumerate(self.columns, 1):
                cell = ws.cell(row=5, column=col_idx, value=self.column_headers.get(col_key, col_key))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            
            # Data rows
            for row_idx, item in enumerate(self.filtered_data, 6):
                for col_idx, col_key in enumerate(self.columns, 1):
                    value = item.get(col_key, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save and open
            temp_dir = tempfile.gettempdir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{self.window_title.replace(' ', '_')}_{timestamp}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            os.startfile(filepath)
            
            messagebox.showinfo("Export Complete", f"Data exported to:\n{filepath}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export to Excel:\n{str(e)}")
    
    def center_window(self):
        """Center the window on screen"""
        self.update_idletasks()
        width = self.winfo_width()
        height = self.winfo_height()
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

# Factory functions
def create_window(parent, title="Window", resize_handles=None, movable=True, location_persistence="none", close_on=None):
    """
    Create a simple custom window
    
    Args:
        parent: Parent window (required)
        title: Window title, or None to hide titlebar completely
        resize_handles: List of ["left", "right", "top", "bottom"] or None
        movable: Whether the window can be moved by dragging (default: True)
        location_persistence: Position persistence mode - "none", "session", or "permanent" (default: "none")
        close_on: List of close methods - ["x_button", "click_outside", "toggle"] (default: ["x_button"])
        
    Returns:
        SimpleWindow instance
    """
    return SimpleWindow(parent, title=title, resize_handles=resize_handles, movable=movable, location_persistence=location_persistence, close_on=close_on)


def create_inventory_window(parent, data, window_config=None):
    """
    Create an inventory view window for displaying tabular data with filtering
    
    Args:
        parent: Parent window (required)
        data: List of dictionaries containing the data to display
        window_config: Configuration dictionary (optional)
        
    Returns:
        InventoryViewWindow instance
    """
    return InventoryViewWindow(parent, data, window_config)


def create_filter_dialog(parent, column_key, column_header, unique_values, current_selection, apply_callback):
    """
    Create a filter dialog for column filtering
    
    Args:
        parent: Parent window (required)
        column_key: Key of the column being filtered
        column_header: Display name of the column
        unique_values: List of unique values for filtering
        current_selection: Currently selected values
        apply_callback: Callback function to apply filter
        
    Returns:
        FilterMenuDialog instance
    """
    return FilterMenuDialog(parent, column_key, column_header, unique_values, current_selection, apply_callback)


def create_data_view_window(parent, data, title="Data View", columns=None, **kwargs):
    """
    Create a data view window for displaying any tabular data with filtering
    
    Args:
        parent: Parent window (required)
        data: List of dictionaries containing the data to display
        title: Window title
        columns: List of column configurations (optional, auto-generated if not provided)
        **kwargs: Additional configuration options for InventoryViewWindow
        
    Returns:
        InventoryViewWindow instance
    """
    window_config = kwargs.copy()
    window_config['title'] = title
    if columns:
        window_config['columns'] = columns
    
    return InventoryViewWindow(parent, data, window_config)


# Utility functions for managing window positions
def clear_session_positions():
    """Clear all session window positions"""
    global _session_window_positions
    _session_window_positions.clear()


def clear_permanent_positions():
    """Clear all permanently saved window positions"""
    try:
        if os.path.exists(_permanent_positions_file):
            os.remove(_permanent_positions_file)
        return True
    except Exception as e:
        return False


def get_saved_positions():
    """Get all saved window positions for debugging/management"""
    positions = {
        "session": _session_window_positions.copy(),
        "permanent": {}
    }
    
    try:
        if os.path.exists(_permanent_positions_file):
            with open(_permanent_positions_file, 'r') as f:
                positions["permanent"] = json.load(f)
    except Exception as e:
        pass
    
    return positions


# Example usage
if __name__ == "__main__":
    # Create root window
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    
    # Create a window that can be resized from left and right
    window = create_window(root, "Test Window", resize_handles=["left", "right"])
    
    # Add some content
    label = tk.Label(window.get_content_frame(), 
                    text="This window can be resized\nfrom the left and right edges",
                    bg="#e8f5e8", fg="#2d5a2d")
    label.pack(pady=20)
    
    # Create a non-movable window as an example
    # fixed_window = create_window(root, "Fixed Window", movable=False)
    
    # Create windows with different persistence modes
    # session_window = create_window(root, "Session Window", location_persistence="session")
    # permanent_window = create_window(root, "Permanent Window", location_persistence="permanent")
    
    # Start the app
    root.mainloop()