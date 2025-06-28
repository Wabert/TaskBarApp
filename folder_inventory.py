# folder_inventory.py (FIXED)
"""
Enhanced Folder Inventory feature for SuiteView Taskbar
Updated to work with new InventoryViewWindow pattern
"""

import os
import threading
import time
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import tempfile
import subprocess
from inventory_view_window import InventoryViewWindow
from explorer_utils import ExplorerDetector

from config import Colors, Fonts, Dimensions
from ui_components import CustomDialog, FormField, WarningDialog, ErrorDialog
from utils import UIUtils

class FolderInventoryDialog(CustomDialog):
    """Enhanced dialog for configuring folder inventory scan with View/Print options"""
    
    def __init__(self, parent):
        super().__init__(parent, "Folder Inventory", width=500, height=400, resizable=True)
        
        self.selected_folder = ""
        self.scan_thread = None
        self.cancel_scan = False
        self.progress_window = None
        
        self.create_form()
        self.add_buttons()
        
        # Auto-populate folder field with topmost Explorer folder (NEW)
        self.auto_populate_folder()
        
        # Set focus on folder field
        self.folder_field.widget.focus_set()
    
    def auto_populate_folder(self):
        """Auto-populate the folder field with the topmost open File Explorer folder"""
        try:
            best_folder = ExplorerDetector.get_best_default_folder()
            if best_folder:
                self.folder_field.set(best_folder)
                self.selected_folder = best_folder
        except Exception as e:
            print(f"Error auto-populating folder: {e}")
            # Fallback to user home directory
            try:
                fallback_folder = os.path.expanduser('~')
                self.folder_field.set(fallback_folder)
                self.selected_folder = fallback_folder
            except:
                pass  # If all else fails, leave empty
        
    def create_form(self):
        """Create the configuration form"""
        # Folder selection using FormField with stacked layout
        folder_container = tk.Frame(self.dialog_content, bg=Colors.LIGHT_GREEN)
        folder_container.pack(fill=tk.X, pady=(0, 10))
        
        self.folder_field = FormField(folder_container, "Folder to scan:", layout='stacked')
        self.folder_field.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        browse_btn = tk.Button(folder_container, text="Browse...", 
                            command=self.browse_folder,
                            bg=Colors.MEDIUM_GREEN, fg=Colors.BLACK,
                            font=Fonts.DIALOG_BUTTON, cursor='hand2')
        browse_btn.pack(side=tk.RIGHT, padx=(5, 0))

        # Depth selection
        self.depth_field = FormField(
            self.dialog_content, "Scan depth:", 
            field_type='combobox',
            values=["1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "Unlimited"],
            width=15
        )
            
        self.depth_field.set("1")
        self.depth_field.widget.configure(state='readonly')
        self.depth_field.pack(fill=tk.X, pady=5)
        
        # Content type selection
        content_frame = tk.Frame(self.dialog_content, bg=Colors.LIGHT_GREEN)
        content_frame.pack(fill=tk.X, pady=10)
        
        tk.Label(content_frame, text="Include:", bg=Colors.LIGHT_GREEN, 
                fg=Colors.BLACK, font=Fonts.DIALOG_LABEL).pack(anchor='w')
        
        self.content_type = tk.StringVar(value="both")
        
        radio_frame = tk.Frame(content_frame, bg=Colors.LIGHT_GREEN)
        radio_frame.pack(fill=tk.X, pady=5)
        
        tk.Radiobutton(radio_frame, text="Files Only", variable=self.content_type, 
                      value="files", bg=Colors.LIGHT_GREEN, fg=Colors.BLACK,
                      font=Fonts.DIALOG_LABEL, activebackground=Colors.MEDIUM_GREEN).pack(side=tk.LEFT, padx=10)
        
        tk.Radiobutton(radio_frame, text="Folders Only", variable=self.content_type, 
                      value="folders", bg=Colors.LIGHT_GREEN, fg=Colors.BLACK,
                      font=Fonts.DIALOG_LABEL, activebackground=Colors.MEDIUM_GREEN).pack(side=tk.LEFT, padx=10)
        
        tk.Radiobutton(radio_frame, text="Both", variable=self.content_type, 
                      value="both", bg=Colors.LIGHT_GREEN, fg=Colors.BLACK,
                      font=Fonts.DIALOG_LABEL, activebackground=Colors.MEDIUM_GREEN).pack(side=tk.LEFT, padx=10)
        
        # Output option selection (NEW)
        output_frame = tk.Frame(self.dialog_content, bg=Colors.LIGHT_GREEN)
        output_frame.pack(fill=tk.X, pady=10)
        
        tk.Label(output_frame, text="Output to:", bg=Colors.LIGHT_GREEN, 
                fg=Colors.BLACK, font=Fonts.DIALOG_LABEL).pack(anchor='w')
        
        self.output_type = tk.StringVar(value="view")  # Default to View
        
        output_radio_frame = tk.Frame(output_frame, bg=Colors.LIGHT_GREEN)
        output_radio_frame.pack(fill=tk.X, pady=5)
        
        tk.Radiobutton(output_radio_frame, text="View (Interactive)", variable=self.output_type, 
                      value="view", bg=Colors.LIGHT_GREEN, fg=Colors.BLACK,
                      font=Fonts.DIALOG_LABEL, activebackground=Colors.MEDIUM_GREEN).pack(side=tk.LEFT, padx=10)
        
        tk.Radiobutton(output_radio_frame, text="Excel File", variable=self.output_type, 
                      value="excel", bg=Colors.LIGHT_GREEN, fg=Colors.BLACK,
                      font=Fonts.DIALOG_LABEL, activebackground=Colors.MEDIUM_GREEN).pack(side=tk.LEFT, padx=10)
        
        # Information text
        info_text = ("View will show results in an interactive grid with filtering.\n"
                    "Excel will create a spreadsheet file.\n"
                    "Large operations (>20,000 items) will show a warning.")
        info_label = tk.Label(self.dialog_content, text=info_text, 
                             bg=Colors.LIGHT_GREEN, fg=Colors.DARK_GREEN,
                             font=(Fonts.DIALOG_LABEL[0], Fonts.DIALOG_LABEL[1]-1),
                             wraplength=450, justify=tk.LEFT)
        info_label.pack(pady=10)
    
    def browse_folder(self):
        """Browse for folder to scan"""
        folder = filedialog.askdirectory(parent=self, title="Select Folder to Inventory")
        if folder:
            self.folder_field.set(folder)
            self.selected_folder = folder
    
    def add_buttons(self):
        """Add Start and Cancel buttons"""
        button_container = tk.Frame(self.button_frame, bg=Colors.LIGHT_GREEN)
        button_container.pack(expand=True)
        
        start_btn = tk.Button(button_container, text="Start Scan", 
                             bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                             command=self.start_scan, width=Dimensions.DIALOG_BUTTON_WIDTH,
                             font=Fonts.DIALOG_BUTTON, relief=tk.RAISED, bd=1,
                             cursor='hand2')
        start_btn.pack(side=tk.LEFT, padx=10)
        
        cancel_btn = tk.Button(button_container, text="Cancel", 
                              bg=Colors.INACTIVE_GRAY, fg=Colors.WHITE,
                              command=self.cancel, width=Dimensions.DIALOG_BUTTON_WIDTH,
                              font=Fonts.DIALOG_BUTTON, relief=tk.RAISED, bd=1)
        cancel_btn.pack(side=tk.LEFT, padx=10)
        
        start_btn.focus_set()
    
    def start_scan(self):
        """Start the folder inventory scan"""
        folder = self.folder_field.get().strip()
        if not folder:
            WarningDialog.show(self, "Invalid Input", "Please select a folder to scan.")
            return
        
        if not os.path.exists(folder):
            ErrorDialog.show(self, "Folder Not Found", f"The selected folder does not exist:\n{folder}")
            return
        
        # Get scan parameters
        depth_str = self.depth_field.get()
        max_depth = None if depth_str == "Unlimited" else int(depth_str)
        content_type = self.content_type.get()
        output_type = self.output_type.get()
        
        # Quick estimation for large operation warning
        if self.should_warn_large_operation(folder, max_depth):
            result = WarningDialog.show(self, "Large Operation Warning", 
                                       "This operation may scan more than 20,000 items and could take several minutes.\n\n"
                                       "Do you want to continue?")
            if not result:
                return
        
        # Start the scan
        self.destroy()
        self.start_inventory_scan(folder, max_depth, content_type, output_type)
    
    def should_warn_large_operation(self, folder, max_depth):
        """Quick check to estimate if operation will be large"""
        try:
            sample_count = 0
            with os.scandir(folder) as entries:
                for entry in entries:
                    sample_count += 1
                    if sample_count >= 1000:
                        break
            
            if sample_count >= 1000:
                return True
            
            if max_depth is None or max_depth > 1:
                with os.scandir(folder) as entries:
                    for entry in entries:
                        if entry.is_dir():
                            return True
            
            return False
        except:
            return False
    
    def start_inventory_scan(self, folder, max_depth, content_type, output_type):
        """Start the actual inventory scan in background thread"""
        self.cancel_scan = False
        
        # Create progress window
        self.progress_window = ProgressWindow(self.parent, self.cancel_scan_operation)
        
        # Start scan in background thread
        self.scan_thread = threading.Thread(
            target=self.perform_scan,
            args=(folder, max_depth, content_type, output_type),
            daemon=True
        )
        self.scan_thread.start()
    
    def cancel_scan_operation(self):
        """Cancel the running scan operation"""
        self.cancel_scan = True
        if self.progress_window:
            self.progress_window.destroy()
    
    def perform_scan(self, folder, max_depth, content_type, output_type):
        """Perform the actual folder scan (runs in background thread)"""
        try:
            scanner = FolderScanner(folder, max_depth, content_type, self.update_progress)
            scanner.cancel_scan = self.cancel_scan  # Pass cancel flag
            inventory_data, error_data = scanner.scan()
            
            if not self.cancel_scan:
                if output_type == "excel":
                    # Create Excel file (existing functionality)
                    excel_creator = ExcelInventoryCreator()
                    excel_creator.create_workbook(inventory_data, error_data, folder)
                else:
                    # Show in interactive view (UPDATED)
                    self.show_inventory_view(inventory_data, error_data, folder, max_depth, content_type)
                
                # Close progress window
                if self.progress_window:
                    self.progress_window.after(0, self.progress_window.destroy)
                
        except Exception as e:
            # Show error in main thread
            if self.progress_window:
                self.progress_window.after(0, lambda: self.show_scan_error(str(e)))
    
    def show_inventory_view(self, inventory_data, error_data, scanned_folder, max_depth, content_type):
        """Show the inventory data in an interactive view (UPDATED FOR NEW PATTERN)"""
        # Create configuration for the new InventoryViewWindow
        window_config = {
            'title': 'Folder Inventory Results',
            'columns': [
                {'key': 'Name', 'header': 'Name', 'width': 200, 'type': 'text'},
                {'key': 'Full Path', 'header': 'Full Path', 'width': 300, 'type': 'text'},
                {'key': 'Type', 'header': 'Type', 'width': 80, 'type': 'text'},
                {'key': 'Size', 'header': 'Size', 'width': 100, 'type': 'text'},
                {'key': 'Modified Date', 'header': 'Modified Date', 'width': 150, 'type': 'date'}
            ],
            'on_item_double_click': self._open_file_or_folder,
            'show_stats': True,
            'allow_export': True,
            'window_width': 1000,
            'window_height': 700,
            'additional_info': {
                'Scanned Folder': scanned_folder,
                'Generated': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Max Depth': max_depth if max_depth else 'Unlimited',
                'Content Type': content_type.title(),
                'Errors': len(error_data) if error_data else 0
            }
        }
        
        # Show the inventory view in the main thread
        if self.progress_window:
            self.progress_window.after(0, lambda: InventoryViewWindow(
                self.parent, inventory_data, window_config))
    
    def _open_file_or_folder(self, item):
        """Handle double-click to open file or folder"""
        try:
            path = item.get('Full Path')
            if path and os.path.exists(path):
                os.startfile(path)
        except Exception as e:
            print(f"Error opening {path}: {e}")
    
    def update_progress(self, count, current_path):
        """Update progress display (called from background thread)"""
        if self.progress_window and not self.cancel_scan:
            self.progress_window.after(0, lambda: self.progress_window.update_progress(count, current_path))
    
    def show_scan_error(self, error_message):
        """Show scan error (called in main thread)"""
        if self.progress_window:
            self.progress_window.destroy()
        ErrorDialog.show(self.parent, "Scan Error", f"An error occurred during scanning:\n{error_message}")


class ProgressWindow(CustomDialog):
    """Progress window for folder scanning"""
    
    def __init__(self, parent, cancel_callback):
        super().__init__(parent, "Scanning Folder", width=450, height=200)
        
        self.cancel_callback = cancel_callback
        
        # Progress info
        self.progress_label = tk.Label(self.dialog_content, text="Starting scan...", 
                                      bg=Colors.LIGHT_GREEN, fg=Colors.BLACK,
                                      font=Fonts.DIALOG_LABEL, wraplength=400)
        self.progress_label.pack(pady=10)
        
        self.count_label = tk.Label(self.dialog_content, text="Items processed: 0", 
                                   bg=Colors.LIGHT_GREEN, fg=Colors.DARK_GREEN,
                                   font=Fonts.DIALOG_LABEL)
        self.count_label.pack(pady=5)
        
        self.path_label = tk.Label(self.dialog_content, text="", 
                                  bg=Colors.LIGHT_GREEN, fg=Colors.DARK_GREEN,
                                  font=(Fonts.DIALOG_LABEL[0], Fonts.DIALOG_LABEL[1]-1),
                                  wraplength=400, anchor='w')
        self.path_label.pack(pady=5, fill=tk.X)
        
        # Cancel button
        cancel_btn = tk.Button(self.button_frame, text="Cancel Scan", 
                              bg=Colors.INACTIVE_GRAY, fg=Colors.WHITE,
                              command=self.cancel_callback, width=12,
                              font=Fonts.DIALOG_BUTTON, relief=tk.RAISED, bd=1)
        cancel_btn.pack()
        
        # Override close button to call cancel
        self.protocol("WM_DELETE_WINDOW", self.cancel_callback)
    
    def update_progress(self, count, current_path):
        """Update progress display"""
        self.count_label.config(text=f"Items processed: {count:,}")
        
        # Truncate long paths
        display_path = current_path
        if len(display_path) > 60:
            display_path = "..." + display_path[-57:]
        
        self.path_label.config(text=f"Current: {display_path}")


class FolderScanner:
    """Core folder scanning logic"""
    
    def __init__(self, root_folder, max_depth, content_type, progress_callback):
        self.root_folder = root_folder
        self.max_depth = max_depth
        self.content_type = content_type
        self.progress_callback = progress_callback
        self.cancel_scan = False
        
        self.inventory_data = []
        self.error_data = []
        self.item_count = 0
    
    def scan(self):
        """Main scan method"""
        self._scan_directory(self.root_folder, 0)
        return self.inventory_data, self.error_data
    
    def _scan_directory(self, directory, current_depth):
        """Recursively scan directory"""
        if self.cancel_scan:
            return
        
        # Check depth limit
        if self.max_depth is not None and current_depth >= self.max_depth:
            return
        
        try:
            with os.scandir(directory) as entries:
                for entry in entries:
                    if self.cancel_scan:
                        return
                    
                    try:
                        self._process_entry(entry, current_depth)
                    except Exception as e:
                        self._log_error(entry.path, str(e))
                    
                    # Update progress every 100 items
                    if self.item_count % 100 == 0:
                        self.progress_callback(self.item_count, entry.path)
        
        except Exception as e:
            self._log_error(directory, str(e))
    
    def _process_entry(self, entry, current_depth):
        """Process a single file/folder entry"""
        is_file = entry.is_file()
        is_dir = entry.is_dir()
        
        # Check if we should include this type
        if self.content_type == "files" and not is_file:
            pass  # Skip folders
        elif self.content_type == "folders" and not is_dir:
            pass  # Skip files
        else:
            # Include this item
            self._add_inventory_item(entry, is_file)
        
        # Recurse into subdirectories
        if is_dir:
            self._scan_directory(entry.path, current_depth + 1)
    
    def _add_inventory_item(self, entry, is_file):
        """Add item to inventory data"""
        try:
            stat_info = entry.stat()
            
            # Get size (only for files)
            if is_file:
                size_bytes = stat_info.st_size
                size_human = self._format_size(size_bytes)
            else:
                size_bytes = 0
                size_human = "N/A"
            
            # Get modified date
            modified_date = datetime.fromtimestamp(stat_info.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            
            item_data = {
                'Name': entry.name,
                'Full Path': entry.path,
                'Type': 'File' if is_file else 'Folder',
                'Size (Bytes)': size_bytes if is_file else '',
                'Size': size_human,
                'Modified Date': modified_date
            }
            
            self.inventory_data.append(item_data)
            self.item_count += 1
            
        except Exception as e:
            self._log_error(entry.path, str(e))
    
    def _format_size(self, bytes_size):
        """Format file size in human readable format"""
        if bytes_size == 0:
            return "0 B"
        
        size_names = ["B", "KB", "MB", "GB", "TB"]
        i = 0
        size = float(bytes_size)
        
        while size >= 1024.0 and i < len(size_names) - 1:
            size /= 1024.0
            i += 1
        
        if i == 0:
            return f"{int(size)} {size_names[i]}"
        else:
            return f"{size:.1f} {size_names[i]}"
    
    def _log_error(self, path, error_message):
        """Log access error"""
        self.error_data.append({
            'Path': path,
            'Error': error_message,
            'Timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })


class ExcelInventoryCreator:
    """Creates Excel workbook with inventory data (existing functionality)"""
    
    def create_workbook(self, inventory_data, error_data, scanned_folder):
        """Create and open Excel workbook"""
        try:
            # Create workbook
            wb = Workbook()
            
            # Create inventory sheet
            ws_inventory = wb.active
            ws_inventory.title = "Inventory"
            self._create_inventory_sheet(ws_inventory, inventory_data, scanned_folder)
            
            # Create errors sheet if there are errors
            if error_data:
                ws_errors = wb.create_sheet("Access Errors")
                self._create_errors_sheet(ws_errors, error_data)
            
            # Save to temp file
            temp_dir = tempfile.gettempdir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Folder_Inventory_{timestamp}.xlsx"
            filepath = os.path.join(temp_dir, filename)
            
            wb.save(filepath)
            
            # Open in Excel
            self._open_excel_file(filepath)
            
        except Exception as e:
            raise Exception(f"Failed to create Excel file: {str(e)}")
    
    def _create_inventory_sheet(self, worksheet, data, scanned_folder):
        """Create the main inventory sheet"""
        # Header
        worksheet['A1'] = f"Folder Inventory: {scanned_folder}"
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        worksheet['A3'] = f"Total Items: {len(data):,}"
        
        # Column headers
        headers = ['Name', 'Full Path', 'Type', 'Size', 'Modified Date']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        
        # Data rows
        for row, item in enumerate(data, 6):
            worksheet.cell(row=row, column=1, value=item['Name'])
            worksheet.cell(row=row, column=2, value=item['Full Path'])
            worksheet.cell(row=row, column=3, value=item['Type'])
            worksheet.cell(row=row, column=4, value=item['Size'])
            worksheet.cell(row=row, column=5, value=item['Modified Date'])
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_errors_sheet(self, worksheet, error_data):
        """Create the errors sheet"""
        # Header
        worksheet['A1'] = "Access Errors"
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet['A2'] = f"Total Errors: {len(error_data)}"
        
        # Column headers
        headers = ['Path', 'Error', 'Timestamp']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
        
        # Error rows
        for row, error in enumerate(error_data, 5):
            worksheet.cell(row=row, column=1, value=error['Path'])
            worksheet.cell(row=row, column=2, value=error['Error'])
            worksheet.cell(row=row, column=3, value=error['Timestamp'])
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _open_excel_file(self, filepath):
        """Open Excel file"""
        try:
            os.startfile(filepath)
        except:
            try:
                subprocess.run(['start', 'excel', filepath], shell=True)
            except:
                messagebox.showinfo("File Created", f"Excel file created at:\n{filepath}")


# Integration function for taskbar.py
def add_folder_inventory_to_taskbar(taskbar_instance):
    """Add Folder Inventory button to the taskbar"""
    
    def show_inventory_dialog():
        """Show the folder inventory dialog"""
        dialog = FolderInventoryDialog(taskbar_instance.root)
        dialog.lift()
        dialog.focus_force()
    
    # Create the inventory button
    inventory_btn = tk.Button(taskbar_instance.main_frame, text="Inventory", 
                             bg=Colors.DARK_GREEN, fg=Colors.WHITE,
                             relief=tk.FLAT, font=Fonts.TASKBAR_BUTTON, 
                             cursor='hand2', activebackground=Colors.HOVER_GREEN, 
                             bd=0, padx=15, command=show_inventory_dialog)
    
    return inventory_btn